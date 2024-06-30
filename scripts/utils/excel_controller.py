"""This module is responsible for the manipulation of workbooks."""
import openpyxl
import openpyxl.styles
import openpyxl.styles.builtins
import xlwings
from copy import copy

class ExcelController:
    """This class is responsible for the manipulation of workbooks."""

    def __init__(self, file_path: str, data_only=False) -> None:
        """
        Initializes a new instance of the ExcelController class.

        Parameters:
            file_path (str): The path to the Excel file.

        Returns:
            None
        """
        self.file_path: str = file_path
        self.workbook : openpyxl.Workbook = self.load_workbook(file_path, data_only)
        self.sheet = self.workbook.active


    def load_workbook(self, file_path: str, data_only=False) -> openpyxl.Workbook:
        """
        Loads an Excel workbook from a file path.

        Args:
            file_path (str): The path to the Excel file.
            data_only (bool, optional): Whether to load the data only or not. Defaults to False.

        Returns:
            openpyxl.Workbook: The loaded workbook.
        """
        if data_only:
            self.parse_data_only(file_path)
        return openpyxl.load_workbook(file_path, data_only=data_only)
    
    def parse_data_only(self, file_path: str):
        """
        Parses the data from an Excel file by opening it with xlwings, saving it, and closing it.

        Args:
            file_path (str): The path to the Excel file.

        Returns:
            None
        """
        app = xlwings.App(visible=False)
        book = app.books.open(file_path)
        book.save()
        book.close()
        app.quit()
    
    def change_sheet(self, sheet_name: str) -> None:
        """
        Changes the active sheet of the workbook.

        Args:
            sheet_name (str): The name of the sheet to be set as active.

        Returns:
            None
        """
        self.sheet = self.workbook[sheet_name]

    def insert_col(self, idx:int) -> None:
        """
        Inserts a column in the workbook.

        Args:
            col (str): The column to be inserted.

        Returns:
            None
        """
        self.sheet.insert_cols(idx)

    def insert_data_into_cell(self, data: int | str | float, cell: str) -> None:
        """
        Inserts data into a cell in the workbook.

        Args:
            data (int | str | float): The data to be inserted into the cell.
            cell (str): The cell address to insert the data.

        Returns:
            None
        """
        # if type(data) == int or type(data) == float:
        #     prev_value = self.sheet[cell].value or 0
        #     self.sheet[cell] = prev_value + data
        # else:
        self.sheet[cell] = data

    def read_from_cell(self, cell:str) -> str:
        """
        Reads data from a cell in the workbook.

        Args:
            cell (str): The cell address to read data from. Should be in the format "A1".

        Returns:
            str: The data read from the cell.
        """
        return self.sheet[cell].value

    def save(self, filename: str) -> None:
        """
        Saves the workbook to a file.

        Args:
            filename (str): The name of the file to save the workbook to.

        Returns:
            None
        """
        self.workbook.save(filename)

    def copy_cell_style_to(self, source_cell, dest_cell:str) -> None:
        self.sheet[dest_cell].fill = copy(self.sheet[source_cell].fill)
        self.sheet[dest_cell].font = copy(self.sheet[source_cell].font)

    def make_cell_accounting(self, cell:str):
        self.sheet[cell].style = "Currency"